import streamlit as st
import streamlit.components.v1 as components 
import openpyxl
from openpyxl.styles import Font, Border, Side
from io import BytesIO
import datetime
import requests
import json
import base64
from fpdf import FPDF
from PIL import Image, ImageOps

# --- 1. IMPOSTAZIONI PAGINA E CSS MODERNO (ADATTIVO DARK/LIGHT MODE) ---
st.set_page_config(page_title="Gestione Note Spese", page_icon="💶", layout="wide")

st.markdown(
    """
    <style>
    /* Bordi input testuali per abbellirli senza rompere il Dark Mode */
    div[data-testid="stTextInput"] input,
    div[data-testid="stNumberInput"] input {
        border-radius: 8px !important;
        border: 1px solid #28a745 !important;
        font-weight: 600;
    }
    
    /* Stile pulsante di Invio (Verde) */
    div[data-testid="stFormSubmitButton"] button {
        background-color: #28a745 !important;
        color: white !important;
        border-radius: 8px !important;
        border: none !important;
        font-weight: bold !important;
        transition: all 0.3s ease;
        width: 100%;
    }
    div[data-testid="stFormSubmitButton"] button:hover {
        background-color: #218838 !important;
        box-shadow: 0 4px 8px rgba(0,0,0,0.15) !important;
        transform: translateY(-2px);
    }

    /* Stile pulsanti di Download (Blu) */
    div[data-testid="stDownloadButton"] button {
        background-color: #007bff !important;
        color: white !important;
        border-radius: 8px !important;
        width: 100%; 
        font-weight: bold;
        transition: all 0.3s ease;
        border: none !important;
    }
    div[data-testid="stDownloadButton"] button:hover {
        background-color: #0056b3 !important;
        box-shadow: 0 4px 8px rgba(0,0,0,0.15) !important;
        transform: translateY(-2px);
    }

    /* Stile pulsante Elimina riga e Svuota tutto (Rossi) */
    button[kind="primary"] {
        background-color: #dc3545 !important;
        color: white !important;
        border-radius: 8px !important;
        border: none !important;
        font-weight: bold;
    }
    button[kind="primary"]:hover {
        background-color: #c82333 !important;
    }
    
    /* Box delle spese: Sfondo semi-trasparente per supportare sia Dark che Light Mode */
    div[data-testid="stHorizontalBlock"] {
        background-color: rgba(128, 128, 128, 0.1); 
        padding: 10px;
        border-radius: 8px;
        border: 1px solid rgba(128, 128, 128, 0.2);
        margin-bottom: 5px;
        align-items: center;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# --- 2. JAVASCRIPT: POTENZIATO PER BLOCCARE INVIO E USARE TAB ---
components.html(
    """
    <script>
    const parentDoc = window.parent.document;
    
    // A. Observer per abbassare la tastiera sulla Data
    const observer = new MutationObserver(() => {
        const inputs = parentDoc.querySelectorAll('div[data-testid="stDateInput"] input');
        inputs.forEach(input => {
            if (input.getAttribute('inputmode') !== 'none') {
                input.setAttribute('inputmode', 'none');
            }
        });
    });
    observer.observe(parentDoc.body, { childList: true, subtree: true });

    // B. Blocca rigorosamente l'Invio nei form Streamlit e simula il TAB
    if (!parentDoc.getElementById('enter-blocker-v2')) {
        const marker = parentDoc.createElement('div');
        marker.id = 'enter-blocker-v2';
        parentDoc.body.appendChild(marker);

        parentDoc.addEventListener('keydown', function(e) {
            if (e.key === 'Enter') {
                const active = parentDoc.activeElement;
                
                // Se premiamo invio dentro un input testo o numero
                if (active && active.tagName === 'INPUT' && active.type !== 'file' && active.type !== 'submit') {
                    
                    // STOP ASSOLUTO ALL'INVIO DEL FORM!
                    e.preventDefault(); 
                    e.stopPropagation();
                    e.stopImmediatePropagation();
                    
                    // Trova tutti gli elementi interattivi (Input testuali/numerici e le tendine Selectbox)
                    const focusable = Array.from(parentDoc.querySelectorAll('input:not([disabled]):not([type="hidden"]), div[data-baseweb="select"]'));
                    
                    let index = focusable.indexOf(active);
                    if (index === -1) {
                        index = focusable.findIndex(el => el.contains(active));
                    }

                    // Se troviamo l'elemento corrente e c'è uno successivo
                    if (index > -1 && index < focusable.length - 1) {
                        const nextEl = focusable[index + 1];
                        
                        // Spostiamo il focus
                        if (nextEl.tagName === 'INPUT') {
                            nextEl.focus();
                        } else {
                            // Se è una tendina (Selectbox), la "clicchiamo" per aprirla e darle il focus
                            nextEl.click();
                        }
                    }
                }
            }
        }, true); // 'true' = intercettiamo il tasto prim'ancora che ci arrivi Streamlit
    }
    </script>
    """,
    height=0, width=0
)

# --- 3. CONFIGURAZIONE CLOUD E CHIAVI ---
try:
    BIN_ID = st.secrets["JSONBIN_ID"]
    API_KEY = st.secrets["JSONBIN_KEY"]
    IMGBB_KEY = st.secrets["IMGBB_KEY"]
except KeyError:
    st.error("⚠️ Attenzione: Mancano alcune chiavi segrete in Streamlit Cloud (Settings > Secrets)!")
    st.stop()

URL_JSONBIN = f"https://api.jsonbin.io/v3/b/{BIN_ID}"
HEADERS = {
    "X-Master-Key": API_KEY,
    "Content-Type": "application/json"
}

# --- 4. FUNZIONI DI MEMORIA E FOTO ---
def salva_spese(lista_spese):
    dati_da_salvare = []
    for spesa in lista_spese:
        spesa_copia = spesa.copy()
        spesa_copia["data"] = spesa_copia["data"].strftime("%Y-%m-%d")
        dati_da_salvare.append(spesa_copia)
        
    payload = {"spese": dati_da_salvare}
    
    try:
        res = requests.put(URL_JSONBIN, json=payload, headers=HEADERS, timeout=10)
        return res.status_code == 200
    except Exception as e:
        st.error(f"⚠️ Errore di connessione a JSONBin: {e}")
        return False

def carica_spese():
    try:
        risposta = requests.get(URL_JSONBIN, headers=HEADERS, timeout=10)
        if risposta.status_code == 200:
            record = risposta.json().get("record", {})
            dati_caricati = record if isinstance(record, list) else record.get("spese", [])
            for spesa in dati_caricati:
                spesa["data"] = datetime.datetime.strptime(spesa["data"], "%Y-%m-%d").date()
            return dati_caricati
    except Exception:
        pass
    return []

def carica_foto_imgbb(foto_bytes):
    url = "https://api.imgbb.com/1/upload"
    foto_b64 = base64.b64encode(foto_bytes).decode('utf-8')
    payload = {"key": IMGBB_KEY, "expiration": 2592000, "image": foto_b64}
    
    try:
        res = requests.post(url, data=payload, timeout=15) 
        if res.status_code == 200:
            return res.json()["data"]["url"]
    except Exception as e:
        st.error(f"⚠️ Errore di connessione durante l'invio della foto: {e}")
    return None

if "spese_settimana" not in st.session_state:
    st.session_state.spese_settimana = carica_spese()

# ==========================================
# --- 5. INTERFACCIA UTENTE (UI) -----------
# ==========================================

st.title("Gestione Nota Spese 📝")

# 🔴🔴🔴 LOGICA DEL PROMEMORIA DEL LUNEDÌ 🔴🔴🔴
oggi = datetime.date.today()
if len(st.session_state.spese_settimana) > 0:
    prima_data = st.session_state.spese_settimana[0]["data"]
    
    # Calcoliamo a che settimana appartengono oggi e le spese in memoria
    anno_oggi, sett_oggi, _ = oggi.isocalendar()
    anno_spesa, sett_spesa, _ = prima_data.isocalendar()
    
    # Verifichiamo se le spese sono "vecchie" (settimana o anno precedente)
    is_settimana_precedente = (anno_oggi > anno_spesa) or (anno_oggi == anno_spesa and sett_oggi > sett_spesa)
    
    # Se oggi è Lunedì (weekday 0) e le spese sono della settimana scorsa
    if oggi.weekday() == 0 and is_settimana_precedente:
        st.error(
            "**🚨 PROMEMORIA IMPORTANTE DI LUNEDÌ!**\n\n"
            "Hai ancora in memoria le spese della settimana scorsa. Ricordati di:\n"
            "1. **Scaricare** il PDF e l'Excel.\n"
            "2. **Inviare** la documentazione.\n"
            "3. **Svuotare la lista** (usando il tasto rosso in fondo alla pagina) prima di inserire le spese di questa nuova settimana!",
            icon="🚨"
        )
        st.markdown("---")


# --- SIDEBAR: Area di inserimento ---
with st.sidebar:
    st.header("➕ Nuova Spesa")
    with st.form("form_spese", clear_on_submit=True):
        # 1. Data
        data_input = st.date_input("Data della spesa", datetime.date.today())
        
        # 2. Motivazione
        motivazione = st.text_input("Motivazione (es. Pranzo Cliente Rossi)")
        
        # 3. Importo
        importo = st.number_input("Importo in Euro (€)", min_value=0.0, step=0.01, format="%.2f", value=None)
        
        # 4. Tipo Spesa
        tipo_spesa = st.selectbox(
            "Seleziona la colonna di destinazione",
            ["Fatture - Carta di Credito Nominale (Colonna H)", "Scontrini - Carta di Credito Nominale (Colonna G)", 
             "Scontrini - Contanti (Colonna C)", "Fatture - Contanti (Colonna D)", "Fatture - Bonifico (Colonna I)"]
        )
        
        # 5. Foto
        foto_scontrino = st.file_uploader("📸 Scatta o allega scontrino", type=["png", "jpg", "jpeg", "heic"])
        
        submit = st.form_submit_button("Aggiungi alla lista")

    if submit:
        if motivazione == "" or importo is None or importo <= 0.0:
            st.warning("⚠️ Inserisci una motivazione e un importo > 0.")
        else:
            foto_url = None
            if foto_scontrino is not None:
                with st.spinner("⏳ Elaborazione foto..."):
                    try:
                        img = Image.open(foto_scontrino)
                        img = ImageOps.exif_transpose(img)
                        if img.mode in ("RGBA", "P"): img = img.convert("RGB")
                        img.thumbnail((1200, 1200))
                        
                        buffer = BytesIO()
                        img.save(buffer, format="JPEG", quality=75) 
                        foto_url = carica_foto_imgbb(buffer.getvalue())
                        
                        if not foto_url:
                            st.error("❌ Errore caricamento online.")
                            st.stop()
                    except Exception as e:
                        st.error(f"❌ Errore elaborazione foto: {e}")
                        st.stop()
            
            nuova_spesa = {
                "data": data_input, "motivazione": motivazione,
                "tipo": tipo_spesa, "importo": importo, "foto_url": foto_url 
            }
            st.session_state.spese_settimana.append(nuova_spesa)
            
            with st.spinner("💾 Salvataggio..."):
                if salva_spese(st.session_state.spese_settimana):
                    st.rerun()

# --- MAIN AREA: Riepilogo e Download ---
if len(st.session_state.spese_settimana) > 0:
    
    totale_settimana = sum(spesa["importo"] for spesa in st.session_state.spese_settimana)
    st.metric(label="💶 Totale Spese Inserite", value=f"{totale_settimana:.2f} €")
    st.markdown("---")
    
    st.subheader("🛒 Dettaglio Spese")
    
    for i, spesa in enumerate(st.session_state.spese_settimana):
        col1, col2, col3, col4 = st.columns([1, 4, 2, 1])
        with col1:
            st.write(spesa['data'].strftime('%d/%m/%Y'))
        with col2:
            icona = " 📸" if spesa.get("foto_url") else ""
            st.write(f"**{spesa['motivazione']}**{icona}")
        with col3:
            st.write(f"**{spesa['importo']:.2f} €**")
        with col4:
            if st.button("❌", key=f"del_btn_{i}", help="Elimina questa spesa"):
                vecchia_lista = st.session_state.spese_settimana.copy()
                st.session_state.spese_settimana.pop(i)
                if salva_spese(st.session_state.spese_settimana):
                    st.rerun()
                else:
                    st.session_state.spese_settimana = vecchia_lista 

    st.markdown("---")
    st.subheader("📥 Genera Documenti")
    
    col_dl1, col_dl2 = st.columns(2)
    
    # -- 1. PDF --
    with col_dl1:
        spese_con_foto = [s for s in st.session_state.spese_settimana if s.get("foto_url")]
        if len(spese_con_foto) > 0:
            if st.button("📄 Prepara PDF Scontrini"):
                with st.spinner("⏳ Preparazione..."):
                    pdf = FPDF(orientation="L", unit="mm", format="A4") 
                    for i in range(0, len(spese_con_foto), 3):
                        pdf.add_page()
                        pdf.set_font("Helvetica", size=10)
                        gruppo_spese = spese_con_foto[i:i+3]
                        x_start, larg_foto, spazio = 10, 85, 10   
                        
                        for idx, sp in enumerate(gruppo_spese):
                            x_pos = x_start + idx * (larg_foto + spazio)
                            pdf.set_xy(x_pos, 10)
                            pdf.cell(w=larg_foto, h=10, text=f"{sp['data'].strftime('%d/%m/%Y')} - {sp['importo']} EUR", align="C")
                            pdf.set_xy(x_pos, 15)
                            pdf.cell(w=larg_foto, h=10, text=sp['motivazione'][:30], align="C")
                            try:
                                req = requests.get(sp["foto_url"], timeout=10)
                                pdf.image(BytesIO(req.content), x=x_pos, y=25, w=larg_foto)
                            except:
                                pdf.set_xy(x_pos, 50)
                                pdf.cell(w=larg_foto, h=10, text="Errore caricamento foto", align="C")

                    num_sett = st.session_state.spese_settimana[0]["data"].isocalendar()[1]
                    st.download_button("⬇️ Scarica PDF", data=bytes(pdf.output()), file_name=f"Scontrini_Sett_{num_sett}.pdf", mime="application/pdf")
        else:
            st.info("Nessuna foto caricata per generare il PDF.")

    # -- 2. EXCEL --
    with col_dl2:
        try:
            workbook = openpyxl.load_workbook("modello_spese.xlsx")
            foglio = workbook.active
            foglio.insert_rows(14, amount=3)
            bordo_sottile = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for r in range(4, 17):
                for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]: foglio[f"{c}{r}"].border = bordo_sottile
            
            prima_data = st.session_state.spese_settimana[0]["data"]
            num_sett = prima_data.isocalendar()[1]
            testo_int = f"COME DA ESTRATTI CONTO: settimana n. {num_sett} anno {prima_data.year}"
            
            for c in range(3, 11): 
                cella = foglio.cell(row=1, column=c)
                if cella.value and "COME DA ESTRATTI CONTO" in str(cella.value):
                    cella.value = testo_int; cella.font = Font(bold=True)
                    break
            else:
                foglio["E1"] = testo_int; foglio["E1"].font = Font(bold=True)

            riga_corr = 4
            for spesa in st.session_state.spese_settimana:
                foglio[f"A{riga_corr}"] = spesa["data"].strftime("%d/%m/%Y")
                foglio[f"B{riga_corr}"] = spesa["motivazione"]
                if "Colonna H" in spesa["tipo"]: foglio[f"H{riga_corr}"] = spesa["importo"]
                elif "Colonna G" in spesa["tipo"]: foglio[f"G{riga_corr}"] = spesa["importo"]
                elif "Colonna C" in spesa["tipo"]: foglio[f"C{riga_corr}"] = spesa["importo"]
                elif "Colonna D" in spesa["tipo"]: foglio[f"D{riga_corr}"] = spesa["importo"]
                elif "Colonna I" in spesa["tipo"]: foglio[f"I{riga_corr}"] = spesa["importo"]
                foglio[f"J{riga_corr}"] = spesa["importo"]
                for c in ["A", "B", "C", "D", "G", "H", "I", "J"]: foglio[f"{c}{riga_corr}"].font = Font(bold=False)
                riga_corr += 1
            
            foglio["J17"] = totale_settimana
            foglio["J17"].font = Font(bold=True)
                
            out_xls = BytesIO()
            workbook.save(out_xls)
            out_xls.seek(0)
            
            st.download_button("⬇️ Scarica Excel", data=out_xls, file_name=f"nota_spese_sett_{num_sett}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except FileNotFoundError:
            st.error("❌ ERRORE: File 'modello_spese.xlsx' non trovato.")

    st.markdown("---")
    
    # -- 3. AZZERA TUTTO --
    with st.expander("⚠️ Opzioni Pericolose"):
        st.warning("Attenzione: Questa azione eliminerà tutte le spese correnti dal database.")
        if st.button("🗑️ Svuota la lista e inizia una nuova settimana", type="primary"):
            with st.spinner("⏳ Svuotamento..."):
                if salva_spese([]):
                    st.session_state.spese_settimana = []
                    st.rerun()
else:
    st.info("👈 Usa il menu a sinistra per inserire la tua prima spesa della settimana!")
